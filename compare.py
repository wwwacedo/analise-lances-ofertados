import os
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configurações
PASTA_LANCES = 'lances'
COR_CYAN = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')

# Dicionários
lances_por_empresa = defaultdict(list)
contagem_por_empresa = defaultdict(int)
arquivos = sorted(os.listdir(PASTA_LANCES))

# Leitura e contagem
for i, arquivo in enumerate(arquivos, start=1):
    if arquivo.endswith('.xlsx'):
        df = pd.read_excel(os.path.join(PASTA_LANCES, arquivo), dtype=str)
        for idx, row in df.iterrows():
            try:
                datahora = datetime.strptime(row['Data/Hora'], '%d/%m/%Y %H:%M:%S')
                empresa = row['Licitante'].strip()
                lances_por_empresa[empresa].append((i, datahora, idx + 2))  # +2: header + 1-indexed Excel
                contagem_por_empresa[empresa] += 1
            except:
                continue

# Encontrar lances simultâneos (mesma empresa, ±1s, em lotes diferentes)
ocorrencias = []

for empresa, lances in lances_por_empresa.items():
    lances.sort(key=lambda x: x[1])
    for i in range(len(lances)):
        lote_i, dt_i, linha_i = lances[i]
        for j in range(i + 1, len(lances)):
            lote_j, dt_j, linha_j = lances[j]
            if abs((dt_i - dt_j).total_seconds()) <= 1 and lote_i != lote_j:
                ocorrencias.append({
                    'empresa': empresa,
                    'momento': min(dt_i, dt_j),
                    'lotes': sorted(set([lote_i, lote_j])),
                    'marcar': [(lote_i, linha_i), (lote_j, linha_j)]
                })

# Eliminar duplicatas
ocorrencias_unicas = []
vistos = set()

for oc in ocorrencias:
    chave = (oc['empresa'], oc['momento'].strftime('%Y-%m-%d %H:%M:%S'), tuple(oc['lotes']))
    if chave not in vistos:
        vistos.add(chave)
        ocorrencias_unicas.append(oc)

# Marcar planilhas
for i, arquivo in enumerate(arquivos, start=1):
    if arquivo.endswith('.xlsx'):
        caminho = os.path.join(PASTA_LANCES, arquivo)
        wb = load_workbook(caminho)
        ws = wb.active

        for oc in ocorrencias_unicas:
            for lote, linha in oc['marcar']:
                if lote == i:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=linha, column=col).fill = COR_CYAN

        wb.save(caminho)
        print(f'Planilha modificada: {arquivo}')

# Relatório de lances simultâneos
print("\nResumo de lances simultâneos por mesma empresa:")
for idx, oc in enumerate(ocorrencias_unicas, start=1):
    data_formatada = oc['momento'].strftime('%d/%m/%Y, %H:%M:%S')
    lotes_str = ', '.join(map(str, oc['lotes']))
    print(f"{idx}. lance em {data_formatada} pela empresa '{oc['empresa']}' - ocorreu nos lotes {lotes_str}")


# Contagem por empresa e por lote
num_lotes = len(arquivos)
contagem_lote_empresa = defaultdict(lambda: [0] * num_lotes)

# Preencher contagem por lote
for empresa, lances in lances_por_empresa.items():
    for lote, _, _ in lances:
        contagem_lote_empresa[empresa][lote - 1] += 1

# Listar todas as empresas que existem nas planilhas (mesmo com 0 lances)
empresas_todas = set()
for i, arquivo in enumerate(arquivos, start=1):
    if arquivo.endswith('.xlsx'):
        df = pd.read_excel(os.path.join(PASTA_LANCES, arquivo), dtype=str)
        empresas_todas.update(df['Licitante'].dropna().str.strip())

# Adicionar empresas que não deram nenhum lance
for empresa in empresas_todas:
    if empresa not in contagem_lote_empresa:
        contagem_lote_empresa[empresa] = [0] * num_lotes

# Exibir o relatório final
print("\nLances por lote (formato: empresa: L1 | L2 | L3):")
for empresa in sorted(contagem_lote_empresa):
    lances_lote = contagem_lote_empresa[empresa]
    lances_str = ' | '.join(map(str, lances_lote))
    print(f"- {empresa}: {lances_str}")

# Empresas participantes por lote
print("\nEmpresas participantes por lote:")

# Inicializa estrutura: lote -> set de empresas
empresas_por_lote = defaultdict(set)

for empresa, lances in lances_por_empresa.items():
    for lote, _, _ in lances:
        empresas_por_lote[lote].add(empresa)

# Imprimir por lote
for lote in range(1, num_lotes + 1):
    print(f"\nLote {lote}:")
    empresas = sorted(empresas_por_lote[lote])
    if not empresas:
        print("  Nenhuma empresa participou deste lote.")
    else:
        for idx, nome in enumerate(empresas, start=1):
            print(f"{idx}. {nome}")
